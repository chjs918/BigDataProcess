#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

scoreList = []
rankList = []
count = 0
row_id = 1;
for row in ws: 
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value*0.3
		sum_v += ws.cell(row = row_id, column = 4).value*0.35
		sum_v += ws.cell(row = row_id, column = 5).value*0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		scoreList.append(sum_v)
		count += 1
	row_id += 1

#grade

sortResult = sorted(scoreList, reverse=True)
#print(sortResult)
for i in scoreList:
	rank = 0
	for j in scoreList:
		if i <= j:
			rank += 1
	rankList.append(rank)
#print(rank)


row_id = 2
for i in rankList:
	#print(i)
	if i <= count * 0.3:
		if i <= count *0.15:	
			ws.cell(row=row_id, column = 8, value="A+") 
		else: ws.cell(row=row_id, column = 8, value="A0") 
	elif i <= count * 0.7:
		if i <= count* 0.35:
			ws.cell(row=row_id, column = 8, value="B+")
		else: ws.cell(row=row_id, column = 8, value="B0")
	else:
		if i <= count * 0.5:
			ws.cell(row=row_id, column = 8, value="C+")
		else: ws.cell(row=row_id, column = 8, value="C0")
	row_id += 1
	


wb.save("student.xlsx")
